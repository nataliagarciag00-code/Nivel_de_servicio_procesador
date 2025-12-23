import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from io import BytesIO
import plotly.express as px

# ==========================
# CONFIGURACIÓN GENERAL
# ==========================
st.set_page_config(
    page_title="Dashboard Nivel de Servicio",
    layout="wide"
)

st.title("📊 Dashboard Profesional – Nivel de Servicio")
st.markdown("Sube los archivos **VF.csv** y **CO.csv** para procesar automáticamente")

# ==========================
# SUBIDA DE ARCHIVOS
# ==========================
col_up1, col_up2 = st.columns(2)

with col_up1:
    vf_file = st.file_uploader("📤 Subir VF.csv", type="csv")

with col_up2:
    co_file = st.file_uploader("📤 Subir CO.csv", type="csv")

# ==========================
# PROCESAMIENTO
# ==========================
if vf_file and co_file:

    # ==========================
    # LECTURA
    # ==========================
    VF = pd.read_csv(vf_file, encoding="latin1", sep=",")
    CO = pd.read_csv(co_file, encoding="latin1", sep=",")

    # ==========================
    # LIMPIEZA DE HORAS
    # ==========================
    def to_time(x):
        try:
            return pd.to_datetime(x).time()
        except:
            return pd.NaT

    for col in ['start_time','start_eta','end_time','end_eta']:
        VF[col] = VF[col].apply(to_time)

    for col in ['H Ini','H Fin']:
        CO[col] = CO[col].apply(to_time)

    def to_delta(t):
        return pd.to_timedelta(str(t)) if t is not pd.NaT else pd.to_timedelta(np.nan)

    VF[['start_time','start_eta','end_time','end_eta']] = VF[
        ['start_time','start_eta','end_time','end_eta']
    ].applymap(to_delta)

    CO[['H Ini','H Fin']] = CO[['H Ini','H Fin']].applymap(to_delta)

    # ==========================
    # CÁLCULOS CO
    # ==========================
    CO['val'] = np.where(
        (CO['V'] == 0) |
        (CO[['SL','SM','SR','SJ','SV','SS','SD']].sum(axis=1) == 0),
        0, 1
    )

    CO['hini_inf'] = CO['H Ini'] - pd.to_timedelta("0:30:00")
    CO.loc[CO['H Ini'] < pd.to_timedelta("0:30:00"), 'hini_inf'] += pd.to_timedelta("1 days")

    CO['hini_sup'] = CO['H Ini'] + pd.to_timedelta("0:30:00")
    CO.loc[CO['hini_sup'] > pd.to_timedelta("1 days"), 'hini_sup'] -= pd.to_timedelta("1 days")

    # ==========================
    # CÁLCULOS VF
    # ==========================
    VF['cod_planta'] = VF['group'].astype(str).str[:3]

    VF['viaje_val'] = np.where(
        (VF['Tipo de Viaje'].isin(['N','V'])) &
        (VF['status'].isin([6,7,8])) &
        (VF['shift'] == 'IN'),
        1, 0
    )

    VF['viaje_val_ad'] = np.where(
        (VF['Tipo de Viaje'].isin(['N','V','A'])) &
        (VF['status'].isin([6,7,8])) &
        (VF['shift'] == 'IN'),
        1, 0
    )

    VF['dif_lle'] = (VF['end_eta'] - VF['end_time']).dt.round("1min").fillna(pd.to_timedelta(0))

    VF['ret_val_lle'] = np.where(
        (VF['viaje_val'] == 1) &
        (VF['record_quality'] == 1) &
        (VF['dif_lle'] >= pd.to_timedelta("0:05:00")),
        1, 0
    )

    def get_hini(row):
        temp = CO[
            (CO['val'] == 1) &
            (CO['ID Servicio'] == row['id_servicio']) &
            (CO['hini_inf'] <= row['start_time']) &
            (CO['hini_sup'] >= row['start_time']) &
            (CO['H Fin'] == row['end_time'])
        ]
        return temp['H Ini'].iloc[0] if len(temp) > 0 else pd.NaT

    VF['h_ini'] = VF.apply(get_hini, axis=1)

    VF['dif_sal'] = (VF['start_eta'] - VF['h_ini']).dt.round("1min").fillna(pd.to_timedelta(0))

    VF['ret_val_sal'] = np.where(
        VF['cod_planta'].str[:2] == "HY",
        0,
        np.where(
            (VF['ret_val_lle'] == 1) &
            (VF['dif_sal'] >= pd.to_timedelta("0:05:00")),
            1, 0
        )
    )

    # ==========================
    # TABLAS FINALES
    # ==========================
    td_vf = VF.groupby('cod_planta').agg(
        viaje_val=('viaje_val','sum'),
        ret_val_lle=('ret_val_lle','sum'),
        ret_val_sal=('ret_val_sal','sum')
    ).reset_index()

    td_vf['%ns'] = (td_vf['viaje_val'] - td_vf['ret_val_lle']) / td_vf['viaje_val']
    td_vf['%ns_sal'] = (td_vf['viaje_val'] - td_vf['ret_val_sal']) / td_vf['viaje_val']

    td_vf_ad = VF.groupby('cod_planta').agg(
        viaje_val_ad=('viaje_val_ad','sum'),
        ret_val_lle=('ret_val_lle','sum'),
        ret_val_sal=('ret_val_sal','sum')
    ).reset_index()

    td_vf_ad['%ns_ad'] = (td_vf_ad['viaje_val_ad'] - td_vf_ad['ret_val_lle']) / td_vf_ad['viaje_val_ad']
    td_vf_ad['%ns_sal_ad'] = (td_vf_ad['viaje_val_ad'] - td_vf_ad['ret_val_sal']) / td_vf_ad['viaje_val_ad']

    # ==========================
    # KPIs
    # ==========================
    st.markdown("## 📌 Indicadores Clave")

    k1, k2, k3 = st.columns(3)
    k1.metric("NS Llegada Promedio", f"{td_vf['%ns'].mean():.2%}")
    k2.metric("NS Salida Promedio", f"{td_vf['%ns_sal'].mean():.2%}")
    k3.metric("Total Viajes", int(td_vf['viaje_val'].sum()))

    # ==========================
    # GRÁFICOS
    # ==========================
    st.markdown("## 📈 Gráficos")

    fig1 = px.bar(
        td_vf,
        x='cod_planta',
        y='%ns',
        title='% Nivel de Servicio – Llegada',
        text='%ns'
    )
    fig1.update_traces(texttemplate='%{text:.2%}', textposition='outside')
    fig1.update_layout(yaxis_tickformat='.0%')
    st.plotly_chart(fig1, use_container_width=True)

    comp = td_vf[['cod_planta','%ns','%ns_sal']] \
        .rename(columns={'%ns':'Llegada','%ns_sal':'Salida'}) \
        .melt(id_vars='cod_planta')

    fig2 = px.bar(
        comp,
        x='cod_planta',
        y='value',
        color='variable',
        barmode='group',
        title='Llegada vs Salida'
    )
    fig2.update_layout(yaxis_tickformat='.0%')
    st.plotly_chart(fig2, use_container_width=True)

    # ==========================
    # ANÁLISIS AUTOMÁTICO
    # ==========================
    mejor = td_vf.loc[td_vf['%ns'].idxmax()]
    peor = td_vf.loc[td_vf['%ns'].idxmin()]

    st.markdown("## 🧠 Análisis Profesional Automático")
    st.write(f"""
    • Mejor planta en llegada: **{mejor['cod_planta']}** ({mejor['%ns']:.2%})  
    • Planta con mayor oportunidad: **{peor['cod_planta']}** ({peor['%ns']:.2%})  
    • Promedio general de NS: **{td_vf['%ns'].mean():.2%}**
    """)

    # ==========================
    # EXPORTAR EXCEL
    # ==========================
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        td_vf_ad.to_excel(writer, sheet_name="td_vf_ad", index=False)
        td_vf.to_excel(writer, sheet_name="td_vf", index=False)

    wb = load_workbook(output)
    for sheet, cols in {'td_vf_ad':['D','F'], 'td_vf':['D','F']}.items():
        ws = wb[sheet]
        for col in cols:
            for row in range(2, ws.max_row + 1):
                ws[f"{col}{row}"].number_format = '0.00%'

    final_output = BytesIO()
    wb.save(final_output)

    st.download_button(
        "📥 Descargar Excel Final",
        final_output.getvalue(),
        file_name="nivel_servicio_25.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
